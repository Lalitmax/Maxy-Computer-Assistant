import openpyxl
import speech_recognition as sr
import pyttsx3
import time

import os
import datetime
date = datetime.datetime.now()
date = str(date)
date = date.split()
date = date[0]
date = date.split('-')
date = date[2]
day = int(date)
eng = pyttsx3.init()


def speak(text):

    eng.say(text)
    eng.runAndWait()


def comaaaaand(name):
    r = sr.Recognizer()
    text = ""  # Initialize text to an empty string

    with sr.Microphone() as mic:
        print(name)
        speak(name)  # Assuming `speak` is defined elsewhere to convert text to speech
        print('listening...')

        r.pause_threshold = 0.5  # Reduced pause threshold for better responsiveness
        try:
            audio = r.listen(mic, timeout=45)  # Listen for audio with a timeout
            print('recognizing...')
            text = r.recognize_google(audio, language="en-in")  # Recognize speech
            text = text.lower()
        except Exception as e:
            print(e)  # Print the error message
            return ""  # Return empty string on exception
    return text  # Return recognized text


# Define file path
base_path = os.path.dirname(__file__)
file_path = os.path.join(base_path, "Excels", "Attendancer.xlsx")

try:
    rm = openpyxl.load_workbook(file_path)
except PermissionError:
    print("Error: The file is currently open. Please close it and try again.")
    exit() 

sh1 = rm["Sheet1"]
sh2 = rm["Sheet2"]

length = sh1.max_row
# for write


# for read
def makeAt():
    li_name = []
    for i in range(2, length+1):
        name = sh1.cell(i, 1).value

        rc = comaaaaand(name)

        if ("present sir" in rc) or ("yes sir" in rc) or ("sir present" in rc) or ("present" in rc):
            print('present')
            sh2.cell(row=i, column=1+day, value="Present")
            rm.save("Excels/Attendancer.xlsx")

        else:
            sh2.cell(row=i, column=1+day, value="Not Present")
            print('absent')
            rm.save("Excels/Attendancer.xlsx")
        name=""

 
