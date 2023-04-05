# Maxy
import datetime

# import speech_recognition as sr
import speech_recognition as sr
import pyttsx3
import smtplib  # for email send

import webbrowser
from googletrans import Translator # translator
import os
import subprocess  # for Run any application
import random
import math
import time
import wikipedia
import pywhatkit as pwt  # for youtube video play
import requests  # for temperature get
from googlesearch import search
import openpyxl  # for exel file open and read
import screen_brightness_control as pct
import sys
import Attendancer as atd
from tkinter import *

wind = Tk()
from PIL import ImageTk, Image
# create date today
date = datetime.datetime.now()
date = str(date)
date = date.split()
date = date[0]
# create random number
r = random.randint(0, 7)


# Create a engine
engine = pyttsx3.init()

# set voice of male or female
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)

# set voice rate
rate = engine.getProperty('rate')
engine.setProperty('rate', 148)

# create current datetime
cur_time = datetime.datetime.now().strftime('%H:%M:%S')
# cur_time2=cur_time = datetime.datetime.now().strftime('%H:%M:%S')

# take only hour
hour1 = int(datetime.datetime.now().hour)
minut1 = int(datetime.datetime.now().minute)







def you(strtext):
    str=strtext.split()
    tt=""
    cnt=0
    check=1
    for i in str:
        if check ==0:
            i=i.capitalize()
            check=1
        tt +=i+" "
        cnt +=1
        if(cnt==5):
            tt +="\n"
            cnt=0
            check=0
    tt=tt.capitalize()
  
    mylabel2.config(text="You: "+tt)

def printtext(str):
    # if var.get() != "":
    #   x = var.get()
    #   mylabel.config(text=x)
    # else:
    str=str.split()
    tt=""
    cnt=0
    check=1
    for i in str:
        if check ==0:
            i=i.capitalize()
            check=1
        tt +=i+" "
        cnt +=1
        if(cnt==5):
            tt +="\n"
            cnt=0
            check=0
    tt=tt.capitalize()
    
    mylabel.config(text=tt.capitalize())

   
    
    
   


def speak(text):
    engine.say(text)
    engine.runAndWait()



# app opener
from AppOpener import open

def appopen(text):

    if ("open" in text) or ('start' in text):
        stext = text.split()
        appname = ""
        if ('open' in stext):
            stext.remove('open')
        if ('application' in stext):
            stext.remove('application')
        if ('start' in stext):
            stext.remove('start')
        if ('hello' in stext):
            stext.remove('hello')

        if ('maxi' in stext):
            stext.remove('maxi')
        if ('the' in stext):
            stext.remove('the')
        # if ('show' in stext):
        #     stext.remove('show')
        if ('maxy' in stext):
            stext.remove('maxy')

        for i in stext:
            appname += i + " "
        p = ''
        s = open(appname)
        p = s

        if s != None:
            p = open(s)
            appname=s

            
        if p ==None:
            speak("opening "+appname)
            printtext("Opened "+appname)
            return 1
            
        else:
            return 0



import psutil # for closing app

def close_app(app_name):
    # returns names of running processes
    running_apps = psutil.process_iter(['pid', 'name'])
    found = False
    for app in running_apps:
        sys_app = app.info.get('name').split('.')[0].lower()

        if sys_app in app_name.split() or app_name in sys_app:
            # returns PID of the given app if found running
            pid = app.info.get('pid')

            # deleting the app if asked app is running.(It raises error for some windows apps)
            try:
                app_pid = psutil.Process(pid)
                app_pid.terminate()
                found = True
            except:
                pass

        else:
            pass
    if not found:
        printtext(app_name+" not found running")
    else:
        printtext(app_name+'('+sys_app+')'+' closed')





def wishme(text):
    printtext(text)
   
    if hour1 >= 0 and hour1 <= 12:
        printtext('good morning sir,\n I am maxy, \nhow may i help you')
        speak('good morning sir, I am maxi, \n  how may i help you')

    elif hour1 >= 12 and hour1 <= 18:
        printtext('good afternoon sir, \nI am maxy,  how may i help you')
        speak('good afternoon sir, I am maxi,  how may i help you')

    else:
        printtext('good evening sir,\n I am maxy,  how may i help you')
        speak('good evening sir, I am maxi,  how may i help you')

def sendmail(email, subject, contant, name):
    from email.message import EmailMessage
    email_id = "Enter your Gmail"
    email_pas = "Enter your password"

    msg = EmailMessage()
    msg["subject"] = subject
    msg["from"] = "Lalit kumar yadav"
    msg["to"] = "To send Gmail"
    msg.set_content(contant)


def lower(t):
    return t.lower()


#  take command for do anything
def command():
    
    r = sr.Recognizer()
    with sr.Microphone() as mic:
        # print('listening...')

        r.pause_threshold = 5
        audio = r.listen(mic, timeout=45)

    try:
        print('recognizing...')
        text = r.recognize_google(audio,language="en-in")
        text = text.lower()
    except Exception as e:
        printtext(e)
        return text
    return text


def doooo(text):
    
    you(text)
    if text in ["exit", "quit", "stop"]:
                exit()
    elif text in ['what is the time', "what's the time", 'is time', 'time now', ]:
        you(text)
        # print()
        if hour1 >= 12:
            speak(str(hour1)+str(minut1)+'PM')
            printtext(str(hour1)+":"+str(minut1)+' PM')
            
        else:
            speak(str(hour1)+str(minut1)+' Am')
            printtext(str(hour1)+":"+str(minut1)+' Am')
            

    elif "open gmail account" in text:
        rm = openpyxl.load_workbook("Chitkara_email.xlsx")

        sh1 = rm["Sheet1"]

        length = sh1.max_row

        # for read
        li_name = []
        for i in range(1, length+1):
            li_name.append(sh1.cell(i, 1).value)
        

        speak("which person you want to send email")
        name = command()
        # print(name)
        speak("tell me subject")
        subject = command()
        # print(subject)
        speak("what do you want to send")
        contant = command()
        
        you("which person you want to \nsend email\n->"+name+"\n"+"tell me subject\n->"+subject+"\n"+"what do you want to \nsend\n->"+contant)

        for i in range(1, length):
            ans = lower(li_name[i])
            if name == ans:
                email = sh1.cell(i+1, 2).value

        sendmail(email, subject, contant, name)
       

    elif ("weather of") in text or "weather in" in text:
        
        
        
        d = text.split()
       
        new_list = [d[i]+" " for i in range(2, len(d))]
        strrequest = ""
        for i in new_list:
            strrequest += i

        city_name =strrequest
        you(strrequest)
        data = requests.get("https://api.openweathermap.org/data/2.5/weather?q=" +
                            city_name+"&appid=882bef32516f8990d82b762cae54604c").json()

        weather = "weather is, "+data["weather"][0]["main"]
        Temp = "temperature is, " + \
            str(int(data["main"]["temp"]-273.5))+" degree Celsius"
        description = 'description is, '+data["weather"][0]['description']
        name = "Name, "+data["name"]+","
        code = "Code is, "+str(data["cod"])
        speak(weather+Temp+description+name+code)
        # you(text)
        printtext(weather+"\n"+Temp+"\n"+description+"\n"+name+"\n"+code)
        you(text)

    elif "temperature in" in text:
       
        
        d = text.split()
       
        new_list = [d[i]+" " for i in range(2, len(d))]
        strrequest = ""
        for i in new_list:
            strrequest += i

        city_name =strrequest
        data = requests.get("https://api.openweathermap.org/data/2.5/weather?q=" +
                            city_name+"&appid=882bef32516f8990d82b762cae54604c").json()

        weather = "weather is, "+data["weather"][0]["main"]
        Temp = "temperature is, " + \
            str(int(data["main"]["temp"]-273.5))+" degree Celsius"
        description = 'description is, '+data["weather"][0]['description']
        name = "Name, "+data["name"]+","
        code = "Code is, "+str(data["cod"])
        
        speak(Temp+name)
        printtext(Temp+"\n"+name)
        you(text)
        


   
    elif ("brightness down"in text) or ("brightness up"in text):

        if(text in "brightness down"):
            brightnessLevel=10
        elif (text in "brightness up"):
            brightnessLevel =90
        pct.set_brightness(brightnessLevel)
        printtext("Brightness set "+str(brightnessLevel)+"%")

        
    elif "brightness" in text :
        # text = text.split()

        brightnessLevel = ""
        for i in text:
            if i >= '0' and i <= '9':
                brightnessLevel += i
        pct.set_brightness(int(brightnessLevel))
        printtext("Brightness set "+str(brightnessLevel)+"%")

    elif text in ["take screenshot","take screenshort","screenshot liya jaye","screenshot liya jay","screenshot lelo","screenshot le lo","screenshot take","capture screenshot","screenshot capture","take a screenshot"]:

        hour = int(datetime.datetime.now().hour)
        minut = int(datetime.datetime.now().minute)
        second = int(datetime.datetime.now().second)

        savefile = date+" "+str(hour)+" "+ str(minut)+ " "+ str(second)
        import pyautogui as pg
        import time
        path = "C:/Screenshot by maxy"
        if(os.path.isdir(path)!=True):
            os.mkdir(path)
        # savefile = random(1,1)
        filename = "C:/Screenshot by maxy/"  + savefile + ".png"
        ss = pg.screenshot(filename)
        ss.show()
        # speak("captured screenshot"+path+"directory")
        printtext("captured screenshot\n"+path)
    elif "in hindi" in text:
        
        text = text.split()
        english = ""
        for i in range(len(text)-2):
            english +=text[i]+" "
        trans = Translator()
        autput = trans.translate(english,dest="hi")
        speak("in hindi")
        printtext(autput.text)
    elif ("in english" in text)or ("ine english" in text):
    
        text = text.split()
        english = ""
        for i in range(len(text)-2):
            english +=text[i]+" "
        trans = Translator()
        autput = trans.translate(english,dest="en")
        speak(autput.text)
        printtext(autput.text)
    elif text in ['open music', 'play music', 'music open', 'music play']:
        os.startfile("wmplayer")
        speak("play music")
        printtext('opened music')
    elif ("what's my name" in text) or ("what is my name" in text):
        a=os.getlogin()
        speak("I think your name is "+a)
        printtext("I think your name is "+a)
        
    elif text in ["lock screen","screen lock","screenlock","screen lock kardo","screen lock kiya jaye","do screen lock","do screenlock"]:
        speak("locked")
        printtext("locked computer")
        subprocess.run("Rundll32.exe user32.dll,LockWorkStation")

    elif ("close firefox" in text) or("firefox close" in text) :
        close_app('firefox')
        speak("closed firefox")
        printtext('closed firefox')
    elif ("close chrome" in text) or ("chrome close" in text):
        close_app("chrome")
        speak("closed chrome")
        printtext('closed chrome')
    elif ("close setting" in text) or ("close settings" in text) or("setting close" in text) or ("settings close" in text):
        close_app("settings")
        speak("closed settings")
        printtext('closed settings')
    elif ("close brave" in text) or ("brave close" in text):
        close_app("brave")
        speak("closed brave")
        printtext('closed brave')
    elif ("close photoshop" in text) or ("photoshop close" in text):
        close_app("photoshop")
        speak("closed photoshop")
        printtext('closed photoshop')
    elif ("close wordpad" in text) or ("wordpad close" in text):
        close_app("wordpad")
        speak("closed wordpad")
        printtext('closed wordpad')
    
    elif ("close calculator" in text) or ("calculator close" in text):
        close_app("calc")
        speak("closed calculator")
        printtext('closed calculator')
    elif ("close notepad" in text) or ("notepad close" in text):
        close_app("notepad")
        speak("closed notepad")
        printtext('closed notepad')
    elif ('close powerpoint'in text) or ('powerpoint close' in text):
        close_app("powerpnt")
        speak("closed powerpoint")
        printtext("closed powerpoint")
    elif ('close whatsapp'in text) or ('whatsapp close' in text):
        close_app("whatsapp")
        speak("closed whatsapp")
        printtext("closed whatsapp")
    elif ('close microsoft edge'in text) or ('microsoft edge close' in text):
        close_app("msedge")
        speak("closed microsoft edge")
        printtext("closed microsoft edge")
    
    elif ('close excel'in text) or ('excel close'in text):
        close_app("excel")
        speak("closed excel")
        printtext("closed excel")
    
    elif ('close mspaint'in text) or ('mspaint close'in text) or('close ms paint' in text) or ('ms paint' in text):
        close_app("mspaint")
        speak("closed mspaint")
        printtext("closed mspaint")
    elif ('close word'in text) or ('word close'in text):
        close_app("word")
        speak("closed word")
        printtext("closed word")
    elif ('close camera'in text) or ('camera close'in text):
        close_app("camera")
        speak("closed camera")
        printtext("closed camera")
    elif ('close cmd'in text) or ('cmd close'in text):
        close_app("cmd")
        speak("closed cmd")
        printtext("closed cmd")
    elif ('close git bash'in text) or ('git bash close'in text):
        close_app("git bash")
        speak("closed git bash")
        printtext("closed git bash")
    elif ('close powershell'in text) or ('powershell close'in text):
        close_app("powershell")
        speak("closed powershell")
        printtext("closed powershell")
        # close_app("code")
    elif ('close visual studio code'in text) or ('visual studio code close'in text) or ('vs code close'in text) or ('close vs code'in text):
        close_app("code")
        speak("closed visual studio code")
        printtext("closed visual studio code")
        # close_app("code")
        
    

    elif ("close all tabs and application and shutdown" in text) or ("close all tabs and shutdown" in text):
        speak("sure, wait some seconds. please don't touch any key until i shutdown computer")
        close_app("firefox")
        close_app("chrome")
        close_app("settings")
        close_app("brave")
        close_app("photoshop")
        close_app("wordpad")
        close_app("calc")
        close_app("notepad")
        close_app("powerpnt")
        close_app("whatsapp")
        close_app("msedge")
        close_app("excel")
        close_app("mspaint")
        close_app("word")
        close_app("osk")
        close_app("camera")
        close_app("screenrec")
        close_app("cmd")
        close_app("git bash")
        close_app("git bash")        
        close_app("powershell")   
        speak("closed")
        subprocess.run("shutdown /s")
        
    elif ("close all the tabs" in text) or ("close all application" in text):
        close_app("firefox")
        close_app("chrome")
        close_app("settings")
        close_app("brave")
        close_app("photoshop")
        close_app("wordpad")
        close_app("calc")
        close_app("notepad")
        close_app("powerpnt")
        close_app("whatsapp")
        close_app("msedge")
        close_app("excel")
        close_app("mspaint")
        close_app("word")
        close_app("osk")
        close_app("camera")
        close_app("screenrec")
        # close_app("code")
        close_app("cmd")
        close_app("git bash")
        close_app("git bash")        
          
        speak("closed")  
        printtext("Closed all tabs")   
    
    elif text in ["wishme", "wish me"]:
       
        wishme(text)

    
    elif text in ["who made you","tum ko kisne banaya","who was created you","who created you"]:
        speak("i am made by lalit max")
        printtext("i am made by lalit max!")
    
    elif text in ["open c drive", "open cdrive", "cdrive open", "c drive open","c:"]:
                # printtext(text)
                speak("opening c drive")
                os.startfile("c:")
                printtext("Opened c drive")
    elif text in ["open d drive", "open ddrive", "ddrive open", "d drive open","c:"]:
                # printtext(text)
                speak("opening c drive")
                os.startfile("d:")
                printtext("Opened c drive")
    elif text in ["show system information"]:
        os.startfile('msinfo32')
        speak("showing system information")
        printtext("Opened system information")

    
    elif ("open" in text) or ("start" in text):
        yesnone=appopen(text)

        if yesnone !=1:
            if text in ["open recycle bin","recycle bin"]:
                speak("opening recycle bin")
                os.startfile("shell:recyclebinfolder")
                printtext("Opened recycle bin")
            
            elif text in ["opens the temporary files folder","open temporary file" "open the temporary files folder", "open temporary files folder","open temporary files folder", "temporary files folder open", "open temporary files", "temporary file open", "open temporary file", "temporary files open","temp","show temporary files","delete temporary files"]:
      
                speak("opening the temporary files folder")
                os.startfile('''temp''')
                printtext("Opened temporary files folder")
            

            elif text in ["open this pc", "this pc open"]:
   
                speak("opening this pc")
                os.startfile('\"')
                printtext("Opened this pc")
  

    elif text in ["check windows version", "windows version check", "check window version", "what is window version","ms-windows-store:","dialer"]:
        speak("checking your windows version")
        os.startfile("winver")



 


    elif "on google" in text:
     
        li = text.split()
        st = ""
        for i in range(len(li)-2):
            st = st+li[i]+" "

        pwt.search(st)

    
    


    elif text in ["kaise ho"]:
        speak("i am fine, and how are you")
        printtext("i am fine, and how are you")

    elif text in ["shutdown", "shut down", "shutdown laptop", "laptop shut down", "laptop shutdown", "shut down laptop", "shutdown computer", "shut down computer", "computer shutdown", "computer shut down"]:
       
        a = subprocess.run("shutdown /s")
        speak("hello user i am going to shuting down your computer, wait some seconds. please don't touch any key until i shutdown computer")
        printtext(" I am gonna shuting down computer don't touch any key")

    elif text in ["restart", "restart laptop", "laptop restart", "computer restart", "restart computer"]:
      
        speak("hello user i am going to shuting down your computer, wait some seconds. please don't touch any key until i restart computer")
        printtext(" I am gonna restart  computer don't touch any key")
        subprocess.run("shutdown /r")


    elif 'how are you' in text:
        speak('i am fine, tell me how may i help you')
        printtext('i am fine.\ntell me how may i help you')
        
        
    elif 'what is your name' in text or "what's your name" in text:
        speak('''well, my name is maxy, i wish that everyone had a nickname as cool as mine, so plz keep small your name  ''')
        printtext('''well, my name's maxy"\ni wish that everyone\nhad a nickname as cool\nas mineso plz keep small and\nsort your name  ''')
        
    elif text in ['are you marry me', "will you marry me"]:
        speak("this is one of things, we'd both have to agree on i'd prefer to keep  our friendship as it is. ")
        printtext("this is one of things \nwe'd both have to agree\non i'd prefer to keep \nour friendship as it is.")
        
    elif text in ['what can you do for me']:
        speak("i can do all the work, which is in my might")
        printtext("i can do all the work \n which is in my might")
        
    elif text in ["do something for me"]:
        speak("Ask me any problem, i will try to solve it for you")
        printtext("Ask me any problem \ni will try to solve it \nfor you")
        
    elif text in ['date', "what's date", "what is date", "date", "what's the date today", "today date", "today's date", "what is the date", "what's the date"]:
        
        speak(date)
        printtext(date)
    elif text in ["tell me some jokes", "tell some jokes","tell me joke" ,"tell me some joke", "kucch joke sunao", "kuchh jokes sunao", 'tell me joke ', 'tell me jokes']:
        # rn = random.randrange(0,2)
        rn = 1
        if rn!=1:
            speak("Air hostess asked lalu Prasad yadav. Sir are you vegetarian or Non vegetarian, Lalu said I am indian. Air hostess said okay, Are you shakahari or mansahari, Lalu said hat sasuri I am Bihari")
            printtext("Air hostess asked lalu \nPrasad yadav. \nSir are you vegetarian or \nNon vegetarian \nLalu said I am indian \nAir hostess said okay, \nAre you shakahari or mansahari \nLalu said hat sasuri I am Bihari")
        else:
            speak('''A tortoise was mugged by a gang of snails, he tried to give a statment to the police, but couldn't, I don't know, it all happened so fast!, he said, looking confused ''')
            printtext('''A tortoise was mugged by \na gang of snails he tried \nto give a statment to the \npolice but couldn't "I don't \nknow, it all happened so fast!\n" he said, looking confused ''')
        
    elif   text in ["man nahi lag raha hai","how i feel good","not liking it"]:
        nm = random(1,10)
        speak("Ok so we think you will have to listen to some song then you will feel like it, well let's play the song for this")
        try:
            pwt.playonyt("Romantic song video enjoy "+str(nm))
            speak('playing')
        except:
            speak("network Error Occurred ")
        printtext("Enjoy...")
    elif "print table of" in text or "table of" in text:

        nu = text.split()
        nu = int(nu[-1])
        strjee = ""
        for i in range(1, 11):
            strjee +=str(i*nu)+" "
        speak("The table of "+str(nu))
        printtext("The table of "+str(nu)+"\n\n"+strjee)

    elif "song of" in text:
        try:
            pwt.playonyt(text)
            speak('playing')
        except:
            speak("network Error Occurred ")
    elif "ka video" in text:
        try:
            pwt.playonyt(text)
            speak('playing')
        except:
            speak("network Error Occurred ")

    elif ("on youtube" in text) or ("on yt" in text):
        try:
            pwt.playonyt(text)
            speak('playing')
            printtext("Playing "+text)
        except:
            speak("network Error Occurred ")
    elif "play video" in text:
        try:
            pwt.playonyt(text)
            speak('playing')
            printtext("playing video for you")
        except:
            speak("network Error Occurred ")
    elif "in binary" in text:
        text = text.split()
        li = []
        num = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                num = int(i)
        binary =bin(num)[2:]
        speak("in binary. "+str(binary))
        printtext(str(num)+"\n= "+str(binary))

    elif "in decimal" in text:
        text = text.split()
        num = 0

        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                num = i
        decimal = int(num,2)

        speak("in decimal. "+str(decimal))
        printtext(str(num)+"\n= "+str(decimal))
    elif ("in hexadecimal" in text) or ("in hex" in text):
        text = text.split()
        li = []
        num = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                num = int(i)
        octal =hex(num)[2:]
        speak("in hexadecimal. "+str(octal))
        printtext(str(num)+"\n= "+str(octal))
    elif ("in octal" in text) or ("in oct" in text):
        text = text.split()
        num = 0

        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                num = int(i)
        decimal = oct(num)[2:]

        speak("in octal. "+str(decimal))
        printtext(str(num)+"\n= "+str(decimal))
    elif ("subtract of" in text) or ("minus" in text) or ("-" in text):
      
        text = text.split()
        li = []
        sum = 0
        t=[]
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
               t.append(float(i))
        a=t[0]
        b=t[1]
        sum = a-b
        speak("the answer is "+str(sum))
        printtext("Sum = "+str(sum))

    elif ("sum of" in text) or ("add" in text) or ("+" in text):
      
        text = text.split()
        li = []
        sum = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                sum += float(i)
        speak("the answer is "+str(sum))
        printtext("Sum = "+str(sum))
        
    elif ("even number " in text) or ("odd number" in text):
       
        text = text.split()
        li = []
        num = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                num = int(i)
        if(num % 2 == 0):
            speak(str(num)+" is a even number")
            printtext(str(num)+" is a even number")
            
        else:
            speak(str(num)+" is a odd number")
            printtext(str(num)+" is a odd number")
            

    elif "area of circle" in text:
        text = text.split()
        li = []
        rad = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                rad = float(i)
        area = 3.14*rad*rad
        speak("The area of circle is "+str(area))
        printtext("The area of circle is "+ str(area))
        

    elif ("multiply" in text) or ("multiple" in text) or ("*" in text):
       
        print()
        text = text.split()
        multp = 1
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                multp *= float(i)
        speak("the answer is "+str(multp))
        printtext("the answer is "+ str(multp))
        
    elif ("divided" in text) or ("/" in text) or ("divide" in text) or (("devide" in text)):
        
        text = text.split()
        li = []
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                li.append(float(i))
        a = li[0]
        b = li[1]
        speak("the answer is "+str(a/b))
        printtext("the answer is "+str( a/b))
        

    elif "how to make" in text:
        try:
            pwt.playonyt(text)
            speak("playing")
        except:
            speak("network Error Occurred ")

    elif text in ["do you know chitkara university"]:
        
        speak(  "yes i know chitkara university, it is best private university in the punjab ")
        printtext("yes i know chitkara university, it is the  best private university in the punjab ")
       
    elif "factorial" in text:
    
        text = text.split()
        li = []
        fact = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                fact = int(i)

        fact = math.factorial(fact)
        speak("The answer is "+str(fact))
        printtext("The answer is "+str(fact))
    elif ("match" in text) or ("cricket" in text) or("vs" in text):

        pwt.search(text.capitalize())
        speak("here are some results")
        printtext("here are some results")
        
  
    elif ("who is" in text) or ("who" in text):
        
        pwt.search(text)

    else:
        printtext("sorry i don't understand")
        speak("sorry i don't understand")

    

if __name__ == '__main__':
    def karo():
                
        speak("Listening start")
        
        text = command()
        doooo(text)
       

    def func():
        x = var.get().lower()
        doooo(x)

    #  for photo set without using your local system photo
    def resource_path(relative_path):
        # Get absolute path to resource, works for dev and for PyInstaller """
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(
        os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)


        


    # for title
    wind.title("Maxy")

#  for icon
    path1 = resource_path("mic.png")
    imj = PhotoImage(file=path1)
    wind.iconphoto(False, imj)

    # // window size

    wind.maxsize(width=1480, height=990)
    wind.minsize(width=1480, height=990)
    # wind.geometry("2000x2000")
    wind.colormapwindows()
   
    wind.config(background='#108cff')
    # label 1
    mylabel = Label(wind,font=("arial",20),bg='#1570CB',justify="left" ,border=1)
    mylabel.place(x=890,y=130)
    # show image
    path2 = resource_path("mic.png")
    my_pic5 = Image.open(path2)

    resize_pic5 = my_pic5.resize((100,100
                                  ), Image.ANTIALIAS)
    new_pic5 = ImageTk.PhotoImage(resize_pic5)
    
    lbl=Label(wind,bg="#108cff").place(x=290,y=60)
    
    # label 2
    mylabel2 = Label(wind,font=("arial",20 ),bg="#1570CB",justify="left",border=1   )
    mylabel2.place(x=60,y=130)


    path3 = resource_path("searchicon.png")
    my_pic = Image.open(path3)

    resize_pic = my_pic.resize((53, 52), Image.ANTIALIAS)
    new_pic = ImageTk.PhotoImage(resize_pic)

    
    # wlcome to maxy
    path4 = resource_path("wctm.png")
    my_pic_wctm = Image.open(path4)

    resize_pic_wctm = my_pic_wctm.resize((388, 124), Image.ANTIALIAS)
    my_pic_wctm = ImageTk.PhotoImage(resize_pic_wctm)

    # Maxy
    mylabel2wctm = Label(wind,text = "MAXY",bg="#108cff",font=("Arial Black",28 ) , fg="white")
    mylabel2wctm.pack(pady=10)

    # welcome
    mylabel2wctm3 = Label(wind,text = "WELCOME",bg="#108cff",font=("Arial Black",13 ) , fg="white")
    mylabel2wctm3.place(x=660,y=75)

    def about():
        pass
    # menu bar
    menuBar= Menu(wind)
    wind.config(menu=menuBar)

    file_manu = Menu(menuBar)
    menuBar.add_cascade(label="About", menu=file_manu,activeforeground="#108cff",activebackground="#108cff")
    file_manu.add_cascade(label="Made by \n Lalit Max", command=about)

    # lebel for exit
    file_manu2 = Menu(menuBar)
    menuBar.add_cascade(label="Close",menu=file_manu2,activeforeground="#108cff",activebackground="#108cff")
    file_manu2.add_cascade(label="Exit",command=wind.quit)
    #  Add gmail
    file_manu3 = Menu(menuBar)
    menuBar.add_cascade(label="Gmail",menu=file_manu3,activeforeground="#108cff",activebackground="#108cff")
    # make attendancer
    def makeatd():
        atd.makeAt()

    file_manu4 = Menu(menuBar)
    menuBar.add_cascade(label="Attendancer", menu=file_manu4,activeforeground="#108cff",activebackground="#108cff")
    file_manu4.add_cascade(label="Add Name Of Student")
    file_manu4.add_cascade(label="Start", command=makeatd)
   
    
    def onReturn(event):
        func()
        ent.delete(0,'end')
   
    var = StringVar()
    ent = Entry(wind, width=40, font=("Arial", 23), textvariable=var)
    ent.bind("<Return>",onReturn)
    ent.place(x=185, y=780)

    btn = Button(wind, image=new_pic,cursor="hand2", command=func,activebackground="#108cff",border=2,bg="white")
    btn.place(x=1219 ,y=780)
   
    btn2 = Button(wind, image=new_pic5,cursor="hand2", command=karo,activebackground="#108cff",bg="#108cff",border=2).place(x=690,y=610)
    
    wind.mainloop()
